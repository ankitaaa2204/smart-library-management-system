from tkinter import messagebox
from openpyxl import Workbook
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.drawing.image import Image
import os


# ---------------- DASHBOARD ----------------
def show_dashboard(connect_db):
    conn = connect_db()
    cursor = conn.cursor()

    cursor.execute("SELECT COUNT(*) FROM books")
    titles = cursor.fetchone()[0]

    cursor.execute("SELECT SUM(quantity) FROM books")
    total_books = cursor.fetchone()[0]

    cursor.execute("""
    SELECT COUNT(*)
    FROM issued_books
    """)
    issued = cursor.fetchone()[0]

    conn.close()

    messagebox.showinfo(
        "Dashboard",
        f"Total Titles: {titles}\n"
        f"Total Books: {total_books}\n"
        f"Issued Books: {issued}"
    )


# ---------------- EXPORT TO EXCEL ----------------
def export_to_excel(connect_db):
    conn = connect_db()
    cursor = conn.cursor()

    wb = Workbook()

    # ================= BOOKS SHEET =================
    ws = wb.active
    ws.title = "Books"

    # ===== LOGO =====
    try:
        logo = Image("logo.png")

        logo.width = 100
        logo.height = 100

        ws.add_image(logo, "A1")

    except:
        pass

    # ===== TITLE =====
    ws.merge_cells("A5:E5")

    title_cell = ws["A5"]

    title_cell.value = "SMART LIBRARY MANAGEMENT REPORT"

    title_cell.font = Font(
        size=16,
        bold=True,
        color="FFFFFF"
    )

    title_cell.fill = PatternFill(
        start_color="4F81BD",
        fill_type="solid"
    )

    title_cell.alignment = Alignment(
        horizontal="center"
    )

    # ===== HEADERS =====
    headers = [
        "Book ID",
        "Title",
        "Author",
        "Category",
        "Quantity"
    ]

    ws.append([])
    ws.append(headers)

    for col in range(1, 6):
        cell = ws.cell(row=7, column=col)

        cell.font = Font(
            bold=True,
            color="FFFFFF"
        )

        cell.fill = PatternFill(
            start_color="4BACC6",
            fill_type="solid"
        )

        cell.alignment = Alignment(
            horizontal="center"
        )

    # ===== BOOK DATA =====
    cursor.execute("SELECT * FROM books")

    books = cursor.fetchall()

    for row in books:
        ws.append(row)

    # ===== AUTO WIDTH =====
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter

        for cell in col:
            if cell.value:
                max_length = max(
                    max_length,
                    len(str(cell.value))
                )

        ws.column_dimensions[col_letter].width = max_length + 3

    # ================= STUDENTS =================
    ws_students = wb.create_sheet("Students")

    ws_students.append([
        "Student ID",
        "Name",
        "Course"
    ])

    cursor.execute("SELECT * FROM students")

    for row in cursor.fetchall():
        ws_students.append(row)

    # ================= ISSUED BOOKS =================
    ws_issued = wb.create_sheet("Issued Books")

    ws_issued.append([
        "Student ID",
        "Book ID",
        "Issue Date",
        "Return Date"
    ])
    cursor.execute("""
        SELECT
            student_id,
            book_id,
            issue_date,
            return_date
       FROM issued_books
    """)

    for row in cursor.fetchall():

        ws_issued.append([
           row[0],  # Student ID
           row[1],  # Book ID
           row[2],  # Issue Date
           row[3]   # Return Date
    ])

    # ================= CHART SHEET =================
    ws_chart = wb.create_sheet("Charts")

    cursor.execute("""
    SELECT category, SUM(quantity)
    FROM books
    GROUP BY category
    """)

    cat_data = cursor.fetchall()

    ws_chart.append(["Category", "Quantity"])

    for row in cat_data:
        ws_chart.append(row)

    # ===== BAR CHART =====
    bar = BarChart()

    bar.title = "Books by Category"

    data = Reference(
        ws_chart,
        min_col=2,
        min_row=1,
        max_row=len(cat_data)+1
    )

    cats = Reference(
        ws_chart,
        min_col=1,
        min_row=2,
        max_row=len(cat_data)+1
    )

    bar.add_data(data, titles_from_data=True)
    bar.set_categories(cats)

    ws_chart.add_chart(bar, "E2")

    # ===== PIE CHART =====
    pie = PieChart()

    pie.title = "Category Distribution"

    pie.add_data(data, titles_from_data=True)
    pie.set_categories(cats)

    ws_chart.add_chart(pie, "E20")

    conn.close()

    # ===== SAVE FILE =====
    from datetime import datetime

    file_path = f"library_report_{datetime.now().strftime('%d%m%Y_%H%M%S')}.xlsx"

    wb.save(file_path)

    os.startfile(file_path)

    messagebox.showinfo(
        "Success",
        "Excel Report Created Successfully!"
    )